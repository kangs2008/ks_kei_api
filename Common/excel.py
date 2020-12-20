from openpyxl import load_workbook
import re

path = r'C:\Users\kangs\Desktop\ks\Datas\dd.xlsx'
class Handle_excel():
    def __init__(self, path):
        self.wb = load_workbook(path)

    def get_sheets(self):
        sheets = self.wb.sheetnames
        return sheets
    def get_sheet_by_name(self, sheet_name):
        sheet = self.wb[sheet_name]
        return sheet
    def get_max_row(self, sheet):
        max_row = sheet.max_row
        return max_row
    def get_max_column(self, sheet):
        max_col = sheet.max_column
        return max_col
    def get_row_value(self, sheet, row):
        columns = sheet.max_column
        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = sheet.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata

    def get_cell_value(self, sheet, row, col):
        value = sheet.cell(row,col).value
        return value
    def get_re_parameter(self, dict_kv, para):
        pattern = r'[{](.*?)[}]'
        para = str(para)
        for key in dict_kv.keys():
            if para in key and para != '':
                para = dict_kv.get(key, key)
                # pass
            else:
                res = re.findall(pattern, str(para))
                if res:
                    for r in res:
                        if r in key:
                            para = para.replace('{' + r + '}', str(dict_kv.get(r, r)))
                    # para = [para.replace('{' + r + '}', str(dict_kv.get(r, r))) for r in enumerate(res) if r == key]
        return para

    def getColumnValuesByTitle(self, sheet, exec_value):
        maxRowNum = self.get_max_row(sheet)
        titleRow = self.get_row_value(sheet, 1)
        execValues = []
        if exec_value in titleRow:
            pos = [i for i, title in enumerate(titleRow) if title == exec_value]

            if len(pos) != 0:
                for rownum in range(2, maxRowNum + 1):
                    value = sheet.cell(rownum, pos[0] +1).value
                    if value is None:
                        value = ''
                    execValues.append(value)

                return execValues, pos
            else:
                print("'exec' is nt define")
        else:
            print('the input exec_value is not in titleRow list.')

    def getExecValuesOfSheet(self, sheet, multi, exec_value=None, exec_type=None):
        maxRowNum = self.get_max_row(sheet)
        columnNum = self.get_max_column(sheet)


        sheetValues = []
        if exec_value is None:
            # sheetValues = getAllValuesOfSheet(sheet)

            pass
        else:
            exec, pos = self.getColumnValuesByTitle(sheet, exec_value)

            for row in range(2, maxRowNum + 1):

                execvalue = exec[row - 2]
                tempdict = {}
                if execvalue.lower() == exec_type:
                    for column in range(1, columnNum + 1):
                        title_value = self.get_cell_value(sheet, 1, column) #title key
                        value = sheet.cell(row, column).value

                        if value is None:
                            value = ''
                        for one in kv:
                            # print('one', one)
                            change_value = self.get_re_parameter(one, value) # config sheet dict
                            # print('change_value', change_value)
                            value = change_value
                            # change_value2 = get_re_parameter('config.cofigation', value)
                            # a=a.replace('\n', '').replace(' ', '')
                            # tmp[title_value] = change_value

                            tempdict[title_value] = change_value

                    sheetValues.append(tempdict)
        return sheetValues
    def getExecKVofSheet(self, sheet, exec_value=None, exec_type=None):
        # row = self.get_max_row(sheet)
        maxRowNum = self.get_max_row(sheet)
        tempdict = {}
        if exec_value is None:
            tempdict = self.getAllKVofSheet(sheet)
        else:
            exec, pos = self.getColumnValuesByTitle(sheet, exec_value)

            for row in range(2, maxRowNum + 1):

                execvalue = exec[row - 2]
                if str(execvalue).lower() == exec_type:
                    key = sheet.cell(row, 1).value
                    value = sheet.cell(row, 2).value
                    if value is None:
                        value = ''
                    tempdict[key] = value

        return tempdict

    def getAllKVofSheet(self, sheet='config'):
        maxRowNum = self.get_max_column(sheet)
        columnNum = self.get_max_column(sheet)

        tempdict = {}
        for row in range(2, maxRowNum + 1):
            key = sheet.cell(row, 1).value
            value = sheet.cell(row, 2).value
            if value is None:
                value = ''
            tempdict[key] = value
        return tempdict

    # def kw(self, *args, **kwargs):
    #     print(kwargs.keys()['user'])
#
if __name__ == '__main__':
    print(Handle_excel(path).get_sheets())
    sheet = Handle_excel(path).get_sheet_by_name('test_接口')
    sh2 = Handle_excel(path).get_sheet_by_name('config')
    Handle_excel(path).getColumnValuesByTitle(sh2, 'exec')
    sh3 = Handle_excel(path).get_sheet_by_name('test_UI')
    Handle_excel(path).getColumnValuesByTitle(sh2, 'exec')
    print('22222222')
    dictkv = Handle_excel(path).getExecKVofSheet(sh2, 'exec', 'y')
    dictkv2 = Handle_excel(path).getExecKVofSheet(sh3, 'exec', 'y')
    print('dictkv', dictkv)
    # print(get_row_value(sheet, '1'))
    print('11111111111')

    kv = [dictkv, dictkv2]


    print(Handle_excel(path).getExecValuesOfSheet(sheet, kv, 'exec', 'y'))
    # print(Handle_excel(path).kw(dictkv))
