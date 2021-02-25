from openpyxl import load_workbook
from openpyxl.styles import Font
import re, os, sys

class Handle_excel():
    def __init__(self, filepath):
        self.wb = load_workbook(filepath)

    def get_sheets(self):
        sheets = self.wb.sheetnames
        return sheets

    def get_sheets_by_rule(self, rule='t_'):
        sheets = self.get_sheets()
        slist = []
        for sheet in sheets:
            if sheet.startswith(rule):
                slist.append(sheet)
        return slist

    def get_sheet_by_name(self, sheet_name):
        try:
            sheet = self.wb[sheet_name]
        except:
            sheet = None
        return sheet
    def get_max_row(self, sheet):
        max_row = sheet.max_row
        return max_row
    def get_max_column(self, sheet):
        max_col = sheet.max_column
        return max_col
    def get_row_value(self, sheet, row):
        columns = sheet.max_column
        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = sheet.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata

    def get_cell_value(self, sheet, row, col):
        value = sheet.cell(row, col).value
        return value
    def get_re_parameter(self, dict_kv, para):
        pattern = r'[$][{](.*?)[}]'
        para = str(para)
        for key in dict_kv.keys():
            res = re.findall(pattern, str(para)) # ${aa}
            if res:
                for r in res:
                    if r in key:
                        para = para.replace('${' + r + '}', str(dict_kv.get(r, r)))
                # para = [para.replace('{' + r + '}', str(dict_kv.get(r, r))) for r in enumerate(res) if r == key]
        return para

    def getColumnValuesByTitle(self, sheet, exec_value):
        maxRowNum = self.get_max_row(sheet)
        titleRow = self.get_row_value(sheet, 1)
        execValues = []
        if exec_value in titleRow:
            pos = [i for i, title in enumerate(titleRow) if title == exec_value]

            if len(pos) != 0:
                for rownum in range(2, maxRowNum + 1):
                    value = sheet.cell(rownum, pos[0] + 1).value
                    if value is None:
                        value = ''
                    execValues.append(value)

                return execValues, pos
            else:
                print("'exec' is nt define")
        else:
            print('the input exec_value is not in titleRow list.')

    def getExecValuesOfSheet(self, sheet, exec_value=None, exec_type=None, multi=None):
        maxRowNum = self.get_max_row(sheet)
        columnNum = self.get_max_column(sheet)

        sheetValues = []
        if exec_value is None:
            # sheetValues = getAllValuesOfSheet(sheet)
            print('???')
            pass
        else:
            exec, pos = self.getColumnValuesByTitle(sheet, exec_value)
            for row in range(2, maxRowNum + 1):

                execvalue = exec[row - 2]
                real_pos = pos[0] + 1
                tempdict = {}
                if execvalue.lower() == exec_type:
                    for column in range(1, columnNum + 1):
                        title_value = self.get_cell_value(sheet, 1, column) # title key
                        value = sheet.cell(row, column).value

                        if value is None:
                            value = ''
                        if multi is not None:
                            for one in multi:
                                change_value = self.get_re_parameter(one, value) # config sheet dict
                                tempdict[title_value] = change_value

                        if column == real_pos:
                            tempdict[title_value] = str(row - 1) # 1 2 3
                    sheetValues.append(tempdict)
        return sheetValues

    def get_re_parameter_sheet(self, dict_kv, para):
        pattern = r'[$][{](.*?)[}]'
        para = str(para)
        for val in dict_kv.values():
            res = re.findall(pattern, str(para)) # ${aa}
            if res:
                for r in res:
                    if r in para:
                        para = para.replace('${' + r + '}', str(val))
        return para

    def getExecKVofSheet(self, sheet, exec_value=None, exec_type=None):
        maxRowNum = self.get_max_row(sheet)
        tempdict = {}
        if exec_value is None:
            tempdict = self.getAllKVofSheet(sheet)
        else:
            exec, pos = self.getColumnValuesByTitle(sheet, exec_value)

            for row in range(2, maxRowNum + 1):

                execvalue = exec[row - 2]
                if str(execvalue).lower() == exec_type:
                    key = sheet.cell(row, 1).value
                    value = sheet.cell(row, 2).value
                    if value is None:
                        value = ''
                    tempdict[key] = value

        return tempdict

    def getAllKVofSheet(self, sheet='config'):
        maxRowNum = self.get_max_column(sheet)
        columnNum = self.get_max_column(sheet)

        tempdict = {}
        for row in range(2, maxRowNum + 1):
            key = sheet.cell(row, 1).value
            value = sheet.cell(row, 2).value
            if value is None:
                value = ''
            tempdict[key] = value
        return tempdict


def load_excel(file_name, sheet_name):
    wb = load_workbook(file_name)
    sheet = wb[sheet_name]
    return wb, sheet

def write_to_excel(sheet, set_value, row_pos, col_pos):
    # 字体格式，颜色和大小
    sheet.cell(int(row_pos) + 1, int(col_pos[0]) + 1).value = set_value
    if set_value == 'FAIL':
        font_s = Font(bold=False, color="FF0000", size=10)
        sheet.cell(int(row_pos) + 1, int(col_pos[0]) + 1).font = font_s
    elif set_value == 'PASS':
        font_s = Font(bold=False, color="00FF00", size=10)
        sheet.cell(int(row_pos) + 1, int(col_pos[0]) + 1).font = font_s
    else:
        pass

def excel_to_save(wb, file_name):
    wb.save(file_name)

def get_file_all_dir(file_dir):
    L = []
    L_name = []
    for root, dirs, files in os.walk(file_dir):
        for file in files:
            if file.endswith('.xlsx') and file.startswith('test_'):
                f = os.path.join(root, file)
                fnew = f.replace('\\', '/')
                L.append(fnew)
                L_name.append(file)
    return L, L_name

def get_file_current_dir(path):
    L = []
    L_name = []
    for file in os.listdir(path):
        if file.endswith('.xlsx') and file.startswith('test_'):
            f = os.path.join(path, file)
            fnew = f.replace('\\', '/')
            L.append(fnew)
            L_name.append(file)
    return L, L_name

def excel_to_case(file_name, no_current_folder=None, sheet_name=None):
    if os.path.isfile(file_name):
        if sys.platform != 'win32':

            fileNameList = file_name.split('/')[-1:]
            file_name = file_name.replace('\\', '/')
            filePathList = [file_name]
        else:
            fileNameList = file_name.split('\\')[-1:]
            file_name = file_name.replace('\\', '/')
            filePathList = [file_name]
    else:
        if no_current_folder == '3':
            filePathList, fileNameList = get_file_all_dir(file_name)
        else:
            filePathList, fileNameList = get_file_current_dir(file_name)

    case_list = []
    for i in range(len(fileNameList)):

        sheet_for_replace = Handle_excel(filePathList[i]).get_sheet_by_name('config')
        if sheet_for_replace:
            dictkv = Handle_excel(filePathList[i]).getExecKVofSheet(sheet_for_replace, 'exec', 'y')
            sheet_rule_list = [dictkv]
        else:
            print(f'1Input sheet name was wrong(not in excel), please check it.{filePathList[i]}')
            sheet_rule_list = None

        if sheet_name:

            sheet_list = Handle_excel(filePathList[i]).get_sheets_by_rule()

            case_kv = {}
            if sheet_name in sheet_list:
                sheet = Handle_excel(filePathList[i]).get_sheet_by_name(sheet_name)
                case_kv[f'sheetname'] = sheet_name
                case_kv[f'file'] = fileNameList[i]
                case_kv[f'filepath'] = filePathList[i]

                stepkv = Handle_excel(filePathList[i]).getExecValuesOfSheet(sheet, 'exec', 'y', sheet_rule_list)
                case_kv[f'filesheet'] = stepkv
                case_list.append(case_kv)
            else:
                print(f'2Input sheet name was wrong(not in excel), please check it.{filePathList[i]}')

        else:
            sheet_list = Handle_excel(filePathList[i]).get_sheets_by_rule()

            for k in range(len(sheet_list)):
                case_kv = {}
                count2 = k + 1
                sheet = Handle_excel(filePathList[i]).get_sheet_by_name(sheet_list[k])
                case_kv[f'sheetname'] = sheet_list[k]
                case_kv[f'file'] = fileNameList[i]
                case_kv[f'filepath'] = filePathList[i]

                stepkv = Handle_excel(filePathList[i]).getExecValuesOfSheet(sheet, 'exec', 'y', sheet_rule_list)
                case_kv[f'filesheet'] = stepkv

                case_list.append(case_kv)
    return case_list

if __name__ == '__main__':
    pass
