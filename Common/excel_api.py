from openpyxl import load_workbook
from openpyxl.styles import Font
import re, os, sys

class Handle_excel():
    def __init__(self, filepath):
        # self.path = filepath
        self.wb = load_workbook(filepath)
        # self.save_f = self.wb.save(filepath)

    # def save_excel(self):
    #     self.save_f

    def get_sheets(self):
        sheets = self.wb.sheetnames
        return sheets

    def get_sheets_by_rule(self, rule='t_'):
        sheets = self.get_sheets()
        slist = []
        for sheet in sheets:
            if sheet.startswith(rule):
                slist.append(sheet)
        return slist

    def get_sheet_by_name(self, sheet_name):
        try:
            sheet = self.wb[sheet_name]
        except:
            sheet = None
        return sheet
    def get_max_row(self, sheet):
        max_row = sheet.max_row
        return max_row
    def get_max_column(self, sheet):
        max_col = sheet.max_column
        return max_col
    def get_row_value(self, sheet, row):
        columns = sheet.max_column
        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = sheet.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata

    def get_cell_value(self, sheet, row, col):
        value = sheet.cell(row, col).value
        return value
    def get_re_parameter(self, dict_kv, para):
        pattern = r'[$][{](.*?)[}]'
        para = str(para)
        for key in dict_kv.keys():
            res = re.findall(pattern, str(para)) # ${aa}
            # print('res---', res, para)
            if res:
                for r in res:
                    if r in key:
                        para = para.replace('${' + r + '}', str(dict_kv.get(r, r)))
                # para = [para.replace('{' + r + '}', str(dict_kv.get(r, r))) for r in enumerate(res) if r == key]
        return para



    def get_re_parameter_bak(self, dict_kv, para):
        pattern = r'[$][{](.*?)[}]'
        para = str(para)
        print(dict_kv)
        for key in dict_kv.keys():
            if para in key and para != '':
                # para = dict_kv.get(key, key)
                pass # 只通过re主动替换
            else:
                res = re.findall(pattern, str(para)) # ${aa}
                print('res---', res, para)
                if res:
                    for r in res:
                        if r in key:
                            para = para.replace('${' + r + '}', str(dict_kv.get(r, r)))
                    # para = [para.replace('{' + r + '}', str(dict_kv.get(r, r))) for r in enumerate(res) if r == key]

        return para

    def getColumnValuesByTitle(self, sheet, exec_value):
        maxRowNum = self.get_max_row(sheet)
        titleRow = self.get_row_value(sheet, 1)
        execValues = []
        if exec_value in titleRow:
            pos = [i for i, title in enumerate(titleRow) if title == exec_value]

            if len(pos) != 0:
                for rownum in range(2, maxRowNum + 1):
                    value = sheet.cell(rownum, pos[0] + 1).value
                    if value is None:
                        value = ''
                    execValues.append(value)

                return execValues, pos
            else:
                print("'exec' is nt define")
        else:
            print('the input exec_value is not in titleRow list.')

    def getExecValuesOfSheet(self, sheet, exec_value=None, exec_type=None, multi=None):
        maxRowNum = self.get_max_row(sheet)
        columnNum = self.get_max_column(sheet)
        print('columnNum', columnNum)

        sheetValues = []
        if exec_value is None:
            # sheetValues = getAllValuesOfSheet(sheet)
            print('???')
            pass
        else:
            exec, pos = self.getColumnValuesByTitle(sheet, exec_value)
            # print('t-接口：', exec, pos)
            for row in range(2, maxRowNum + 1):

                execvalue = exec[row - 2]
                real_pos = pos[0] + 1
                tempdict = {}
                if execvalue.lower() == exec_type:
                    for column in range(1, columnNum + 1):
                        title_value = self.get_cell_value(sheet, 1, column) # title key
                        value = sheet.cell(row, column).value

                        if value is None:
                            value = ''
                        if multi is not None:
                            for one in multi:
                                change_value = self.get_re_parameter(one, value) # config sheet dict
                                tempdict[title_value] = change_value

                        if column == real_pos:
                            # tempdict[title_value] = value + str(row - 1) y1 y2
                            tempdict[title_value] = str(row - 1) # 1 2 3
                    sheetValues.append(tempdict)
        return sheetValues

    def get_re_parameter_sheet(self, dict_kv, para):
        pattern = r'[$][{](.*?)[}]'
        para = str(para)
        for val in dict_kv.values():
            res = re.findall(pattern, str(para)) # ${aa}
            print('res---', res, para, val)
            if res:
                for r in res:
                    if r in para:
                        para = para.replace('${' + r + '}', str(val))
        return para







    def getExecKVofSheet(self, sheet, exec_value=None, exec_type=None):
        # row = self.get_max_row(sheet)
        maxRowNum = self.get_max_row(sheet)
        tempdict = {}
        if exec_value is None:
            tempdict = self.getAllKVofSheet(sheet)
        else:
            exec, pos = self.getColumnValuesByTitle(sheet, exec_value)
            # print('exec, pos', exec, pos)
            col_num = ''
            for row in range(2, maxRowNum + 1):

                execvalue = exec[row - 2]
                if str(execvalue).lower() == exec_type:
                    key = sheet.cell(row, 1).value
                    value = sheet.cell(row, 2).value
                    if value is None:
                        value = ''
                    # tempdict[f'col_num{row}'] = value
                    tempdict[key] = value
        # print(tempdict)
        return tempdict

    def getAllKVofSheet(self, sheet='config'):
        maxRowNum = self.get_max_column(sheet)
        columnNum = self.get_max_column(sheet)

        tempdict = {}
        for row in range(2, maxRowNum + 1):
            key = sheet.cell(row, 1).value
            value = sheet.cell(row, 2).value
            if value is None:
                value = ''
            tempdict[key] = value
        return tempdict


def load_excel(file_name, sheet_name):
    # file_name = file_name.replace('\\', '/')
    wb = load_workbook(file_name)
    sheet = wb[sheet_name]
    print('load')
    return wb, sheet

def write_to_excel(sheet, set_value, row_pos):

    # exec, pos = Handle_excel(file_name).getColumnValuesByTitle(sheet, 'return_code')
    # col_pos = pos[0] + 1

    # 字体格式，颜色和大小
    # font_pass = Font(size=16, bold=True, color="00FF00")
    font_false = Font(bold=True, color="FF0000")


    col_pos = 9
    sheet.cell(int(row_pos) + 1, int(col_pos)).font = font_false

    sheet.cell(int(row_pos) + 1, int(col_pos)).value = set_value


def write_to_excel3(sheet, set_value, row_pos, col_pos):

    # exec, pos = Handle_excel(file_name).getColumnValuesByTitle(sheet, 'return_code')
    # col_pos = int(col_pos[0]) + 1

    # 字体格式，颜色和大小
    #
    sheet.cell(int(row_pos) + 1, int(col_pos[0]) + 1).value = set_value
    if set_value == 'FAIL':
        font_s = Font(bold=True, color="FF0000")
        sheet.cell(int(row_pos) + 1, int(col_pos[0]) + 1).font = font_s
    elif set_value == 'PASS':
        font_s = Font(bold=True, color="00FF00")
        sheet.cell(int(row_pos) + 1, int(col_pos[0]) + 1).font = font_s
    else:
        pass


def excel_to_save(wb, file_name):
    wb.save(file_name)











def get_file_all_dir(file_dir):
    L = []
    L_name = []
    for root, dirs, files in os.walk(file_dir):
        for file in files:
            if file.endswith('.xlsx') and file.startswith('test_'):
                f = os.path.join(root, file)
                fnew = f.replace('\\', '/')
                L.append(fnew)
                L_name.append(file)
    return L, L_name


def get_file_current_dir(path):
    L = []
    L_name = []
    for file in os.listdir(path):
        # print('file', file)
        if file.endswith('.xlsx') and file.startswith('test_'):
            f = os.path.join(path, file)
            fnew = f.replace('\\', '/')
            L.append(fnew)
            L_name.append(file)
    return L, L_name

def excel_to_case(file_name, no_current_folder=None, sheet_name=None):
    if os.path.isfile(file_name):
        if sys.platform != 'win32':

            fileNameList = file_name.split('/')[-1:]
            file_name = file_name.replace('\\', '/')
            filePathList = [file_name]
        else:
            fileNameList = file_name.split('\\')[-1:]
            file_name = file_name.replace('\\', '/')
            filePathList = [file_name]
    else:
        if no_current_folder:
            filePathList, fileNameList = get_file_all_dir(file_name)
        else:
            filePathList, fileNameList = get_file_current_dir(file_name)

    case_list = []
    for i in range(len(fileNameList)):

        sheet_for_replace = Handle_excel(filePathList[i]).get_sheet_by_name('config')
        if sheet_for_replace:
            dictkv = Handle_excel(filePathList[i]).getExecKVofSheet(sheet_for_replace, 'exec', 'y')
            sheet_rule_list = [dictkv]
        else:
            print(f'1Input sheet name was wrong(not in excel), please check it.{filePathList[i]}')
            sheet_rule_list = None

        count = i + 1
        if sheet_name:

            sheet_list = Handle_excel(filePathList[i]).get_sheets_by_rule()

            case_kv = {}
            if sheet_name in sheet_list:
                sheet = Handle_excel(filePathList[i]).get_sheet_by_name(sheet_name)
                case_kv[f'sheetname{count}'] = sheet_name
                case_kv[f'file{count}'] = fileNameList[i]
                case_kv[f'filepath{count}'] = filePathList[i]

                stepkv = Handle_excel(filePathList[i]).getExecValuesOfSheet(sheet, 'exec', 'y', sheet_rule_list)
                case_kv[f'filesheet{count}'] = stepkv
                case_list.append(case_kv)
            else:
                print(f'2Input sheet name was wrong(not in excel), please check it.{filePathList[i]}')

        else:
            sheet_list = Handle_excel(filePathList[i]).get_sheets_by_rule()

            for k in range(len(sheet_list)):
                case_kv = {}
                count2 = k + 1
                sheet = Handle_excel(filePathList[i]).get_sheet_by_name(sheet_list[k])
                case_kv[f'sheetname{count2}'] = sheet_list[k]
                case_kv[f'file{count2}'] = fileNameList[i]
                case_kv[f'filepath{count2}'] = filePathList[i]

                stepkv = Handle_excel(filePathList[i]).getExecValuesOfSheet(sheet, 'exec', 'y', sheet_rule_list)
                case_kv[f'filesheet{count2}'] = stepkv

                case_list.append(case_kv)
    return case_list


















if __name__ == '__main__':
    path = r'D:\desk20201127\ks\Datas'
    file = r'D:\desk20201127\ks\Report\2020-12-09_21-08-17\test_apidata_report.xlsx'
    sheet_name = 't_接'

    sheet = Handle_excel(r'D:\desk20201127\ks\Datas\test_apidata.xlsx').get_sheet_by_name('t_接')
    # exec, pos = Handle_excel(file).getColumnValuesByTitle(sheet, 'exec')
    # print('aaaaa:', exec, pos)
    # print(excel_to_case(path, None, None))


    # wb, sheet = load_excel2(r'D:\desk20201127\ks\Report\2020-12-09_21-08-17\test_apidata_report.xlsx', 't_接')
    # write_excel2(wb, sheet, 'set_value', '9')
    # write_to_save2(wb, file)
    # exec, pos1 = Handle_excel(r'D:\desk20201127\ks\Datas\test_apidata.xlsx').getColumnValuesByTitle(sheet, 'return_values')
    # exec, pos2 = Handle_excel(r'D:\desk20201127\ks\Datas\test_apidata.xlsx').getColumnValuesByTitle(sheet, 'return_code')
    # print(pos1)
    # print(pos2)

    # print(pos)
    # print(write_to_excel(file, 't_接', 'valuu', '9'))





    # print(get_file_current_dir(path2))
    # print(get_file_all_dir(path2))

    # sheet_for_replace = Handle_excel(path).get_sheet_by_name('config')
    # dictkv = Handle_excel(path).getExecKVofSheet(sheet_for_replace, 'exec', 'y')
    # kv_list = [dictkv]


    # [{'write1': 'w'}, {'write2': 'w'}, {'write3': 'w'}, {'write4': 'w'}, {'write5': 'w'}, {'write6': 'w'}, {'write7': 'w'}]
    # print(excel_to_case(path, None, ''))
    # print(excel_to_case(path, 1, None))


    path = r'D:\desk20201127\ks\Datas\test_apidata.xlsx'
    #
    # print(Handle_excel(path).get_sheets())
    # sheet = Handle_excel(path).get_sheet_by_name('t_接口')
    sh2 = Handle_excel(path).get_sheet_by_name('config')
    # Handle_excel(path).getColumnValuesByTitle(sh2, 'exec')
    # # sh3 = Handle_excel(path).get_sheet_by_name('test_UI')
    # # Handle_excel(path).getColumnValuesByTitle(sh2, 'exec')
    # print('22222222')
    dictkv = Handle_excel(path).getExecKVofSheet(sh2, 'exec', 'y')
    # # dictkv2 = Handle_excel(path).getExecKVofSheet(sh3, 'exec', 'y')
    # print('dictkv', dictkv)
    # # print(get_row_value(sheet, '1'))
    # print('11111111111')
    #
    kv = [dictkv]
    #
    #
    print(kv)
    print(Handle_excel(path).getExecValuesOfSheet(sheet, 'exec', 'y', kv))
    # # print(Handle_excel(path).kw(dictkv))
